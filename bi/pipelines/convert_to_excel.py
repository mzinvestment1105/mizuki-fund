"""screening_master parquet → Excel（列の一部削除・日本語ヘッダ・スクリーニングシート）"""

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path

import pandas as pd

try:
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    _HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAVE_OPENPYXL = False
    Table = None  # type: ignore[misc, assignment]
    TableStyleInfo = None  # type: ignore[misc, assignment]

INPUT_PATH = Path("..") / "outputs" / "screening_master.parquet"
OUTPUT_PATH = Path("..") / "outputs" / "screening_master.xlsx"

DATA_SHEET = "データ"
COND_SHEET = "スクリーニング条件"
RESULT_SHEET = "スクリーニング結果"
TABLE_NAME = "ScrMaster"


def _jp_headers_weekly_short_vol_val() -> dict[str, str]:
    """信用売り8週・機関空売り8週・出来高/売買代金ブロック8本の日本語ヘッダ（買い週次と同じ命名規則）。"""
    out: dict[str, str] = {}
    for i in range(1, 9):
        m = f"{i:02d}"
        if i == 1:
            wk_suf = "_最古"
        elif i == 8:
            wk_suf = "_直近"
        else:
            wk_suf = ""
        out[f"ShortMargin_WkSeq{m}"] = f"信用売り残_週次{m}{wk_suf}"
        out[f"ShortSale_WkSeq{m}"] = f"機関空売り株数_週次{m}{wk_suf}"
        if i == 1:
            blk_suf = "_最古"
        elif i == 8:
            blk_suf = "_直近"
        else:
            blk_suf = ""
        out[f"VolAvg5d_BlkSeq{m}"] = f"出来高5日平均_ブロック{m}{blk_suf}"
        out[f"ValAvg5d_BlkSeq{m}"] = f"売買代金5日平均_ブロック{m}{blk_suf}"
    return out


DROP_COLS = [
    "DiscretionaryInvestmentContractorName",
    "ShortPositionsToSharesOutstandingRatio",
    "FiscalQuarter",
]

JP_HEADERS: dict[str, str] = {
    "Code": "銘柄コード",
    "CompanyName": "銘柄名",
    "MarketCodeName": "市場",
    "Sector17CodeName": "セクター17",
    "Sector33CodeName": "セクター33",
    "Close": "終値",
    "MarketCap": "時価総額",
    "YFinanceMarketCap": "時価総額_Yahoo",
    "YFinanceSharesOutstanding": "発行済株式数_Yahoo",
    "NetSales_PriorYear_Actual": "売上高_昨年通期実績",
    "NetSales_LatestYear_Actual": "売上高_今年通期実績",
    "NetSales_NextYear_Forecast": "売上高_来年通期予想",
    "OperatingProfit_PriorYear_Actual": "営業利益_昨年通期実績",
    "OperatingProfit_LatestYear_Actual": "営業利益_今年通期実績",
    "OperatingProfit_NextYear_Forecast": "営業利益_来年通期予想",
    "Profit_PriorYear_Actual": "最終益_昨年通期実績",
    "Profit_LatestYear_Actual": "最終益_今年通期実績",
    "Profit_NextYear_Forecast": "最終益_来年通期予想",
    "NetSales_TwoYearsPrior_Actual": "売上高_一昨年通期実績",
    "OperatingProfit_TwoYearsPrior_Actual": "営業利益_一昨年通期実績",
    "Profit_TwoYearsPrior_Actual": "最終益_一昨年通期実績",
    "CashAndEquivalents_LatestFY": "現金及び現金同等物_直近期末",
    "Equity_LatestFY": "純資産額_直近期末",
    "PER_Trailing": "PER_実績ベース",
    "PBR_Trailing": "PBR_実績ベース",
    "ROE_LatestYear": "ROE_実績ベース",
    "EquityToAssetRatio": "自己資本比率",
    "NumberOfIssuedAndOutstandingSharesAtTheEndOfFiscalYearIncludingTreasuryStock": "発行株式総数",
    "ShortMarginTradeVolume": "信用売り残",
    "LongMarginTradeVolume": "信用買い残",
    "LongMargin_WkSeq01": "信用買い残_週次01_最古",
    "LongMargin_WkSeq02": "信用買い残_週次02",
    "LongMargin_WkSeq03": "信用買い残_週次03",
    "LongMargin_WkSeq04": "信用買い残_週次04",
    "LongMargin_WkSeq05": "信用買い残_週次05",
    "LongMargin_WkSeq06": "信用買い残_週次06",
    "LongMargin_WkSeq07": "信用買い残_週次07",
    "LongMargin_WkSeq08": "信用買い残_週次08_直近",
    "AvgDailyVolume5d": "出来高_5日平均",
    "AvgDailyValue5d": "売買代金_5日平均",
    "ShortPositionsInSharesNumber": "空売り残高（株数）",
    "AnnouncementDate": "決算発表予定日",
    "FiscalYear": "会計年度",
    "YFinance_Supplemented": "YF補完",
    "ETLRunId": "ETL実行ID",
    "ETLStartedAtUTC": "ETL開始時刻(UTC)",
    "ETLStartedAtJST": "ETL開始時刻(JST)",
}
JP_HEADERS.update(_jp_headers_weekly_short_vol_val())

# スクリーニング用派生（parquet には含めず Excel のみ）。
# 信用は playbook 準拠: 買残/発行株・買残/出来高を主とする。
DERIVED_JP_HEADERS: dict[str, str] = {
    "Scr_LongMargin_to_SharesOutstanding": "信用買い-発行済比率",
    "Scr_LongMargin_to_AvgVol5d": "信用買い-出来高倍率",
    "Scr_InstShort_to_Mcap": "機関空売り_時価総額比",
    "Scr_Sales_CAGR2y": "売上高_CAGR2年",
    "Scr_Sales_Y1": "売上高_成長率Y1",
    "Scr_Sales_Y2": "売上高_成長率Y2",
    "Scr_OP_CAGR2y": "営業利益_CAGR2年",
    "Scr_OP_Y1": "営業利益_成長率Y1",
    "Scr_OP_Y2": "営業利益_成長率Y2",
    "Scr_NI_CAGR2y": "最終益_CAGR2年",
    "Scr_NI_Y1": "最終益_成長率Y1",
    "Scr_NI_Y2": "最終益_成長率Y2",
    "Scr_Sales_FcstGrowth": "売上高_予想対実績伸び率",
    "Scr_OP_FcstGrowth": "営業利益_予想対実績伸び率",
    "Scr_NI_FcstGrowth": "最終益_予想対実績伸び率",
    "Scr_Cash_to_Mcap": "現金同等物_時価総額比",
}

# スクリーニング条件シートの行順（Data シートの日本語列名と一致）
SCREENING_TABLE_COLUMNS: list[tuple[str, str]] = [
    ("時価総額", "円。下限のみ例:  large cap 向けに 300000000000 など"),
    ("PER_実績ベース", "実績PER（倍）"),
    ("PBR_実績ベース", "実績PBR（倍）"),
    ("ROE_実績ベース", "今期ROE（%）"),
    ("自己資本比率", "小数で入力（例 0.35 = 35%）。データと同じ基準"),
    ("信用買い-発行済比率", "買残÷期末発行済株数（playbook ①）"),
    ("信用買い-出来高倍率", "買残÷5日平均出来高＝解消日数目安（playbook ②）"),
    ("機関空売り_時価総額比", "空売り残株×終値÷時価総額"),
    ("現金同等物_時価総額比", "現金同等物÷時価総額"),
    ("売上高_CAGR2年", "一昨年→今年実績の2年CAGR=(今年/一昨年)^0.5-1。正の実績のみ"),
    ("売上高_成長率Y1", "一昨年→昨年の単年成長率。正の実績のみ"),
    ("売上高_成長率Y2", "昨年→今年の単年成長率。正の実績のみ"),
    ("営業利益_CAGR2年", "同上（営業利益）"),
    ("営業利益_成長率Y1", "一昨年→昨年の単年成長率（営業利益）"),
    ("営業利益_成長率Y2", "昨年→今年の単年成長率（営業利益）"),
    ("最終益_CAGR2年", "同上（最終益）"),
    ("最終益_成長率Y1", "一昨年→昨年の単年成長率（最終益）"),
    ("最終益_成長率Y2", "昨年→今年の単年成長率（最終益）"),
    ("売上高_予想対実績伸び率", "来年予想÷今年実績-1（予想欠損行は計算されません）"),
    ("営業利益_予想対実績伸び率", "同上"),
    ("最終益_予想対実績伸び率", "同上"),
]

HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF000000", bgColor="FF000000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _numeric_col(df: pd.DataFrame, name: str) -> pd.Series:
    """欠損列時も行インデックス揃った Float64 列を返す（.get + to_numeric(None) のスカラー化を防ぐ）。"""
    if name not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")
    return pd.to_numeric(df[name], errors="coerce")


def _add_screening_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    """スクリーニング用の派生列を英語キーで追加（続けて JP へリネーム）。"""
    out = df.copy()
    close = _numeric_col(out, "Close")
    mcap = _numeric_col(out, "MarketCap")
    lm = _numeric_col(out, "LongMarginTradeVolume")
    av5 = _numeric_col(out, "AvgDailyVolume5d")
    inst = _numeric_col(out, "ShortPositionsInSharesNumber")
    cash = _numeric_col(out, "CashAndEquivalents_LatestFY")

    inst_yen = inst * close
    sh_out = _numeric_col(
        out,
        "NumberOfIssuedAndOutstandingSharesAtTheEndOfFiscalYearIncludingTreasuryStock",
    )
    out["Scr_LongMargin_to_SharesOutstanding"] = lm / sh_out
    out["Scr_LongMargin_to_AvgVol5d"] = lm / av5
    out["Scr_InstShort_to_Mcap"] = inst_yen / mcap
    out["Scr_Cash_to_Mcap"] = cash / mcap

    def _cagr2(a0: pd.Series, a1: pd.Series) -> pd.Series:
        s0 = pd.to_numeric(a0, errors="coerce")
        s1 = pd.to_numeric(a1, errors="coerce")
        ratio = s1 / s0
        ok = s0.notna() & s1.notna() & (s0 > 0) & (s1 > 0)
        r = pd.Series(pd.NA, index=out.index, dtype="Float64")
        r.loc[ok] = ratio[ok].pow(0.5) - 1.0
        return r

    out["Scr_Sales_CAGR2y"] = _cagr2(
        _numeric_col(out, "NetSales_TwoYearsPrior_Actual"),
        _numeric_col(out, "NetSales_LatestYear_Actual"),
    )
    out["Scr_OP_CAGR2y"] = _cagr2(
        _numeric_col(out, "OperatingProfit_TwoYearsPrior_Actual"),
        _numeric_col(out, "OperatingProfit_LatestYear_Actual"),
    )
    out["Scr_NI_CAGR2y"] = _cagr2(
        _numeric_col(out, "Profit_TwoYearsPrior_Actual"),
        _numeric_col(out, "Profit_LatestYear_Actual"),
    )

    def _yoy(a0: pd.Series, a1: pd.Series) -> pd.Series:
        """単年成長率: a1/a0 - 1。両方正の実績のみ計算。"""
        s0 = pd.to_numeric(a0, errors="coerce")
        s1 = pd.to_numeric(a1, errors="coerce")
        ok = s0.notna() & s1.notna() & (s0 > 0) & (s1 > 0)
        r = pd.Series(pd.NA, index=out.index, dtype="Float64")
        r.loc[ok] = (s1[ok] / s0[ok]) - 1.0
        return r

    # 売上高 Y1（一昨年→昨年）・Y2（昨年→今年）
    out["Scr_Sales_Y1"] = _yoy(
        _numeric_col(out, "NetSales_TwoYearsPrior_Actual"),
        _numeric_col(out, "NetSales_PriorYear_Actual"),
    )
    out["Scr_Sales_Y2"] = _yoy(
        _numeric_col(out, "NetSales_PriorYear_Actual"),
        _numeric_col(out, "NetSales_LatestYear_Actual"),
    )
    # 営業利益 Y1・Y2
    out["Scr_OP_Y1"] = _yoy(
        _numeric_col(out, "OperatingProfit_TwoYearsPrior_Actual"),
        _numeric_col(out, "OperatingProfit_PriorYear_Actual"),
    )
    out["Scr_OP_Y2"] = _yoy(
        _numeric_col(out, "OperatingProfit_PriorYear_Actual"),
        _numeric_col(out, "OperatingProfit_LatestYear_Actual"),
    )
    # 最終益 Y1・Y2
    out["Scr_NI_Y1"] = _yoy(
        _numeric_col(out, "Profit_TwoYearsPrior_Actual"),
        _numeric_col(out, "Profit_PriorYear_Actual"),
    )
    out["Scr_NI_Y2"] = _yoy(
        _numeric_col(out, "Profit_PriorYear_Actual"),
        _numeric_col(out, "Profit_LatestYear_Actual"),
    )

    def _fcst_growth(f: pd.Series, a: pd.Series) -> pd.Series:
        ff = pd.to_numeric(f, errors="coerce")
        aa = pd.to_numeric(a, errors="coerce")
        ok = ff.notna() & aa.notna() & (aa != 0)
        r = pd.Series(pd.NA, index=out.index, dtype="Float64")
        r.loc[ok] = (ff[ok] / aa[ok]) - 1.0
        return r

    out["Scr_Sales_FcstGrowth"] = _fcst_growth(
        _numeric_col(out, "NetSales_NextYear_Forecast"),
        _numeric_col(out, "NetSales_LatestYear_Actual"),
    )
    out["Scr_OP_FcstGrowth"] = _fcst_growth(
        _numeric_col(out, "OperatingProfit_NextYear_Forecast"),
        _numeric_col(out, "OperatingProfit_LatestYear_Actual"),
    )
    out["Scr_NI_FcstGrowth"] = _fcst_growth(
        _numeric_col(out, "Profit_NextYear_Forecast"),
        _numeric_col(out, "Profit_LatestYear_Actual"),
    )
    return out


# Excel 表示用（科学記数法・#### 緩和）
_NUM_FMT_INT = "#,##0"
_NUM_FMT_FLOAT = "#,##0.00"
_NUM_FMT_PERCENT = "0.00%"
_MIN_WIDTH_BY_JP_HEADER: dict[str, float] = {
    "決算発表予定日": 14,
    "会計年度": 12,
    "銘柄名": 28,
}
_MIN_WIDTH_NUMERIC = 14
_MIN_WIDTH_MONEY = 16
_MAX_COL_WIDTH = 55


def _estimate_display_chars(v: object, *, prefer_float: bool = False) -> int:
    if v is None or v == "":
        return 0
    if isinstance(v, bool):
        return 5
    if isinstance(v, (datetime, date)):
        return 12
    if isinstance(v, str):
        return len(v)
    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            if fv != fv:
                return 0
            if abs(fv) >= 1e15:
                return 14
            if prefer_float or (isinstance(v, float) and abs(fv - round(fv)) > 1e-6):
                return len(f"{fv:,.2f}")
            return len(f"{int(round(fv)):,}")
        except (OverflowError, ValueError):
            return len(str(v))
    return len(str(v))


def _equity_ratio_as_excel_fraction(raw: object) -> float | None:
    try:
        fv = float(raw)
    except (TypeError, ValueError):
        return None
    if fv != fv:
        return None
    if fv > 1.0 + 1e-6:
        return fv / 100.0
    return fv


def _estimate_percent_display_chars(v: object) -> int:
    fr = _equity_ratio_as_excel_fraction(v)
    if fr is None:
        return 0
    return max(len(f"{fr * 100:.2f}%"), 7)


def _style_header_row(ws, header_row: int = 1) -> None:
    for col_idx in range(1, ws.max_column + 1):
        c = ws.cell(header_row, col_idx)
        if c.value is not None:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN


def _apply_excel_display_formats_workbook(path: Path, *, header_row: int = 1) -> None:
    """データシートに数値書式・列幅。ヘッダは黒背景・白太字。"""
    if not _HAVE_OPENPYXL:
        return
    wb = load_workbook(path)
    if DATA_SHEET not in wb.sheetnames:
        wb.save(path)
        return
    ws = wb[DATA_SHEET]
    if ws.max_row < header_row or ws.max_column < 1:
        wb.save(path)
        return

    _style_header_row(ws, header_row)
    ws.freeze_panes = f"A{header_row + 1}"

    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col_idx)
        if cell.value is not None:
            headers[str(cell.value)] = col_idx

    money_jp = {
        "終値",
        "時価総額",
        "時価総額_Yahoo",
        "発行済株式数_Yahoo",
        "売上高_昨年通期実績",
        "売上高_今年通期実績",
        "売上高_来年通期予想",
        "営業利益_昨年通期実績",
        "営業利益_今年通期実績",
        "営業利益_来年通期予想",
        "最終益_昨年通期実績",
        "最終益_今年通期実績",
        "最終益_来年通期予想",
        "売上高_一昨年通期実績",
        "営業利益_一昨年通期実績",
        "最終益_一昨年通期実績",
        "現金及び現金同等物_直近期末",
        "純資産額_直近期末",
        "発行株式総数",
        "信用売り残",
        "信用買い残",
        "信用買い残_週次01_最古",
        "信用買い残_週次02",
        "信用買い残_週次03",
        "信用買い残_週次04",
        "信用買い残_週次05",
        "信用買い残_週次06",
        "信用買い残_週次07",
        "信用買い残_週次08_直近",
        "空売り残高（株数）",
        "出来高_5日平均",
        "売買代金_5日平均",
    }
    money_jp |= set(_jp_headers_weekly_short_vol_val().values())
    ratio_metric_jp = {
        "PER_実績ベース",
        "PBR_実績ベース",
        "ROE_実績ベース",
    }
    screening_ratio_jp = {
        "機関空売り_時価総額比",
        "現金同等物_時価総額比",
    }
    screening_long_margin_pct_jp = {
        "信用買い-発行済比率",
        "信用買い-出来高倍率",
    }
    screening_rate_as_pct_jp = {
        "売上高_CAGR2年",
        "営業利益_CAGR2年",
        "最終益_CAGR2年",
        "売上高_成長率Y1",
        "売上高_成長率Y2",
        "営業利益_成長率Y1",
        "営業利益_成長率Y2",
        "最終益_成長率Y1",
        "最終益_成長率Y2",
        "売上高_予想対実績伸び率",
        "営業利益_予想対実績伸び率",
        "最終益_予想対実績伸び率",
    }

    for jp, cidx in headers.items():
        if jp in money_jp:
            fmt = _NUM_FMT_INT
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, cidx)
                if cell.value is not None and cell.value != "":
                    cell.number_format = fmt
        elif jp in screening_rate_as_pct_jp:
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, cidx)
                if cell.value is not None and cell.value != "":
                    cell.number_format = _NUM_FMT_PERCENT
        elif jp in screening_long_margin_pct_jp:
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, cidx)
                if cell.value is not None and cell.value != "":
                    cell.number_format = _NUM_FMT_PERCENT
        elif jp in ratio_metric_jp or jp in screening_ratio_jp:
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, cidx)
                if cell.value is not None and cell.value != "":
                    cell.number_format = _NUM_FMT_FLOAT
        elif jp == "自己資本比率":
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, cidx)
                if cell.value is not None and cell.value != "":
                    fr = _equity_ratio_as_excel_fraction(cell.value)
                    if fr is not None:
                        cell.value = fr
                    cell.number_format = _NUM_FMT_PERCENT

    _skip_auto_numeric_headers = {"銘柄コード", "ETL実行ID"}

    for cidx in range(1, ws.max_column + 1):
        header_val = ws.cell(header_row, cidx).value
        header_s = str(header_val) if header_val is not None else ""
        if header_s in _skip_auto_numeric_headers:
            continue
        if "日" in header_s or "Date" in header_s or "時刻" in header_s:
            continue
        if (
            header_s in money_jp
            or header_s in ratio_metric_jp
        or header_s in screening_ratio_jp
        or header_s in screening_rate_as_pct_jp
        or header_s == "自己資本比率"
        ):
            continue

        nonnull = 0
        numeric = 0
        any_float = False
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, cidx).value
            if v is None or v == "":
                continue
            nonnull += 1
            if isinstance(v, bool):
                continue
            if isinstance(v, (datetime, date)):
                numeric = 0
                nonnull = 0
                break
            if isinstance(v, (int, float)):
                numeric += 1
                if isinstance(v, float) and abs(v - int(v)) > 1e-9:
                    any_float = True
            elif isinstance(v, str) and v != "予想無し":
                numeric = 0
                nonnull = 0
                break

        if nonnull == 0 or numeric / nonnull < 0.8:
            continue

        fmt = _NUM_FMT_FLOAT if any_float else _NUM_FMT_INT
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(r, cidx)
            if cell.value is not None and cell.value != "":
                if cell.number_format not in ("General", "0", "0.0", "0.00"):
                    continue
                cell.number_format = fmt

    for jp, cidx in headers.items():
        letter = get_column_letter(cidx)
        hdr_w = len(jp) + 3
        max_cell = hdr_w
        floor = _MIN_WIDTH_BY_JP_HEADER.get(jp, 10)
        if jp in money_jp:
            floor = max(floor, _MIN_WIDTH_MONEY)
        elif jp in ratio_metric_jp or jp in screening_ratio_jp or jp == "自己資本比率":
            floor = max(floor, _MIN_WIDTH_NUMERIC)
        elif jp in screening_rate_as_pct_jp:
            floor = max(floor, 12.0)

        prefer_float = jp in ratio_metric_jp or jp in screening_ratio_jp
        is_eq_ratio = jp == "自己資本比率"
        is_rate_pct = jp in screening_rate_as_pct_jp

        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, cidx).value
            if v is None or v == "":
                continue
            if is_eq_ratio:
                wch = _estimate_percent_display_chars(v)
            elif is_rate_pct:
                try:
                    fv = float(v)
                    wch = max(len(f"{fv * 100:.2f}%"), 9) if fv == fv else 0
                except (TypeError, ValueError):
                    wch = _estimate_display_chars(v, prefer_float=True)
            else:
                wch = _estimate_display_chars(v, prefer_float=prefer_float)
            if wch > max_cell:
                max_cell = wch

        target = min(_MAX_COL_WIDTH, max(max_cell + 2, floor, hdr_w))
        ws.column_dimensions[letter].width = float(target)

    _append_table_and_screening_sheets(wb, ws, header_row)
    wb.save(path)


def _append_table_and_screening_sheets(wb, data_ws, header_row: int) -> None:
    """Excel テーブル登録と条件・結果シート（Excel 365 の FILTER 想定）。"""
    if not _HAVE_OPENPYXL or Table is None:
        return
    max_r = data_ws.max_row
    max_c = data_ws.max_column
    if max_r < 2 or max_c < 1:
        return
    last_letter = get_column_letter(max_c)
    ref = f"A{header_row}:{last_letter}{max_r}"

    for _tname in list(data_ws.tables):
        del data_ws.tables[_tname]

    tab = Table(displayName=TABLE_NAME, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data_ws.add_table(tab)

    if COND_SHEET in wb.sheetnames:
        del wb[COND_SHEET]
    if RESULT_SHEET in wb.sheetnames:
        del wb[RESULT_SHEET]

    cws = wb.create_sheet(COND_SHEET)
    cws.merge_cells("A1:D1")
    c1 = cws["A1"]
    c1.value = (
        "↓ 各指標の下限・上限を入力してください（空欄＝その側の制約なし）。"
        " Microsoft 365 の Excel で「スクリーニング結果」を開いてください。"
    )
    c1.alignment = Alignment(wrap_text=True, vertical="top")
    c1.font = Font(bold=True)
    cws.row_dimensions[1].height = 36

    for col, title in enumerate(["指標（データ列名）", "下限", "上限", "メモ"], start=1):
        cell = cws.cell(2, col, title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    cws.column_dimensions["A"].width = 32
    cws.column_dimensions["B"].width = 14
    cws.column_dimensions["C"].width = 14
    cws.column_dimensions["D"].width = 52

    start_data = 3
    for i, (col_jp, memo) in enumerate(SCREENING_TABLE_COLUMNS):
        r = start_data + i
        cws.cell(r, 1, col_jp)
        cws.cell(r, 2, None)
        cws.cell(r, 3, None)
        cws.cell(r, 4, memo)

    rws = wb.create_sheet(RESULT_SHEET)
    rws["A1"] = "条件をすべて満たす行（データシートの全列を表示）"
    rws["A1"].font = Font(bold=True)
    formula_parts: list[str] = []
    cond_ref = f"'{COND_SHEET}'"
    for i, (col_jp, _) in enumerate(SCREENING_TABLE_COLUMNS):
        r = start_data + i
        esc = col_jp.replace("'", "''")
        formula_parts.append(
            f"(IF(ISBLANK({cond_ref}!$B${r}),TRUE,{TABLE_NAME}[{esc}]>={cond_ref}!$B${r}))"
        )
        formula_parts.append(
            f"(IF(ISBLANK({cond_ref}!$C${r}),TRUE,{TABLE_NAME}[{esc}]<={cond_ref}!$C${r}))"
        )
    combined = "*".join(formula_parts)
    formula = (
        f'=IFERROR(FILTER({TABLE_NAME},{combined},"該当がありません"),'
        f'"FILTER未対応: Microsoft 365 の Excel で開いてください")'
    )
    rws["A2"] = formula
    rws.column_dimensions["A"].width = 24

    # 条件シートのヘッダ行の高さ
    cws.row_dimensions[2].height = 28


def parquet_to_excel(
    inp: Path, outp: Path, *, max_rows: int | None = None
) -> tuple[Path, pd.DataFrame]:
    if not inp.exists():
        raise FileNotFoundError(f"input parquet not found: {inp}")

    df = pd.read_parquet(inp)
    if max_rows is not None and max_rows > 0:
        df = df.head(int(max_rows))
    outp.parent.mkdir(parents=True, exist_ok=True)

    _wk8_blk8_optional = [
        *[f"LongMargin_WkSeq{i:02d}" for i in range(1, 9)],
        *[f"ShortMargin_WkSeq{i:02d}" for i in range(1, 9)],
        *[f"ShortSale_WkSeq{i:02d}" for i in range(1, 9)],
        *[f"VolAvg5d_BlkSeq{i:02d}" for i in range(1, 9)],
        *[f"ValAvg5d_BlkSeq{i:02d}" for i in range(1, 9)],
        "AvgDailyValue5d",
    ]
    for _c in _wk8_blk8_optional:
        if _c not in df.columns:
            df[_c] = pd.NA

    for _col in ("ShortMarginTradeVolume", "LongMarginTradeVolume"):
        if _col in df.columns:
            df[_col] = pd.to_numeric(df[_col], errors="coerce")
    if "AvgDailyVolume5d" in df.columns:
        df["AvgDailyVolume5d"] = pd.to_numeric(df["AvgDailyVolume5d"], errors="coerce")
    if "AvgDailyValue5d" in df.columns:
        df["AvgDailyValue5d"] = pd.to_numeric(df["AvgDailyValue5d"], errors="coerce")
    _wk_blk_prefixes = (
        "LongMargin_WkSeq",
        "ShortMargin_WkSeq",
        "ShortSale_WkSeq",
        "VolAvg5d_BlkSeq",
        "ValAvg5d_BlkSeq",
    )
    for _c in list(df.columns):
        if any(_c.startswith(p) for p in _wk_blk_prefixes):
            df[_c] = pd.to_numeric(df[_c], errors="coerce")

    drop = [c for c in DROP_COLS if c in df.columns]
    if drop:
        df = df.drop(columns=drop)

    df = _add_screening_derived_columns(df)

    rename = {k: v for k, v in JP_HEADERS.items() if k in df.columns}
    rename.update({k: v for k, v in DERIVED_JP_HEADERS.items() if k in df.columns})
    df = df.rename(columns=rename)

    _forecast_cols_jp = ["売上高_来年通期予想", "営業利益_来年通期予想", "最終益_来年通期予想"]
    for _fc in _forecast_cols_jp:
        if _fc in df.columns:
            df[_fc] = df[_fc].where(df[_fc].notna(), other="予想無し")

    try:
        with pd.ExcelWriter(outp, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=DATA_SHEET, index=False)
    except PermissionError as e:
        raise PermissionError(
            f"Excel ファイルに書けません（開いたままの Excel を閉じてください）: {outp}"
        ) from e
    try:
        _apply_excel_display_formats_workbook(outp)
    except Exception:
        pass
    return outp, df


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--input", type=Path, default=INPUT_PATH)
    p.add_argument("--output", type=Path, default=OUTPUT_PATH)
    args = p.parse_args()
    out, _ = parquet_to_excel(args.input, args.output)
    print(out)


if __name__ == "__main__":
    main()
